from openpyxl import load_workbook

# Carrega o arquivo xlsx (excel) que tem os dados do estoque
workbook = load_workbook(filename='C:/Users/saogv/PycharmProjects/Controle de Estoque/estoque.xlsx')
sheet = workbook.active

####Com esse código é possível adicionar o novo item ao final da planilha, com o código, nome e quantidade
codigo = 111
nome = 'dddd'
quantidade = 1123
sheet.append([codigo, nome, quantidade])

# Código do produto que vai ter o valor alterado, esse código precisa acompanhar os 3 codigos abaixo
produto = 12

####Com esse codigo é possível atualizar a quantidade de um item
for row in sheet.iter_rows(min_row=2):
    codigo = row[0].value
    quantidade = row[2].value

    # Verifica se o nome do produto na planilha é igual ao nome do produto que deseja atualizar
    if codigo == produto:
        nova_quantidade = 185  # Nova quantidade que deseja atualizar
        row[2].value = nova_quantidade  # Atualiza a quantidade na planilha
        break  # Sai do loop após encontrar o produto
else:
    print(f'O produto {produto} não foi encontrado na planilha.')

####com esse codigo é possivel retirar uma quantidade de um item
for row in sheet.iter_rows(min_row=2):
    codigo = row[0].value
    quantidade = row[2].value

    # Verifica se o nome do produto na planilha é igual ao nome do produto que deseja atualizar
    if codigo == produto:
        quantidade_retirada = 162  # quantidade que vai ser retirada
        row[2].value -= quantidade_retirada # Atualiza a quantidade na planilha
        break  # Sai do loop após encontrar o produto
else:
    print(f'O produto {produto} não foi encontrado na planilha.')


####com esse codigo é possivel acrescentar uma quantidade a um item no caso de uma devolução
for row in sheet.iter_rows(min_row=2):
    codigo = row[0].value
    quantidade = row[2].value

    # Verifica se o nome do produto na planilha é igual ao nome do produto que deseja atualizar
    if codigo == produto:
        quantidade_devolvida = 1  # quantidade que vai ser retirada
        row[2].value += quantidade_devolvida # Atualiza a quantidade na planilha
        break  # Sai do loop após encontrar o produto
else:
    print(f'O produto {produto} não foi encontrado na planilha.')



# Salva as alterações no arquivo XLSX
workbook.save(filename='C:/Users/saogv/PycharmProjects/Controle de Estoque/estoque.xlsx')


#### mostra todo o estoque
tab_cod = 8
tab_nome = 35
tab_qtd = 10

# Itera pelas linhas da planilha a partir da segunda linha
print(f'{"Código".ljust(tab_cod)}{"|"}{"Nome".ljust(tab_nome)}{"|"}{"Quantidade".ljust(tab_qtd)}')
for row in sheet.iter_rows(min_row=2):
    codigo = row[0].value
    nome = row[1].value
    quantidade = row[2].value

    # Exibe os dados na tela
    print(f'{str(codigo).ljust(tab_cod)}{"|"}{str(nome).ljust(tab_nome)}{"|"}{str(quantidade).ljust(tab_qtd)}')


#### com esse código é possivel exluir um item da planilha

# Procura pelo nome do item a ser excluído
item_a_excluir = 222
linha_a_excluir = None

for linha in sheet.iter_rows():
    item = linha[0].value
    if item == item_a_excluir:
        linha_a_excluir = linha
        print("O ítem", item_a_excluir, "foi excluído com sucesso!")
        break

# Remove a linha se ela foi encontrada
if linha_a_excluir is not None:
    sheet.delete_rows(linha_a_excluir[0].row)



# Salva as alterações no arquivo
workbook.save(filename='C:/Users/saogv/PycharmProjects/Controle de Estoque/estoque.xlsx')





#### abaixo tem uma opção feita sem o xlsx, porem tudo é mais manual e os dados não são salvos. Para ativar retirar os """
"""
#Aqui tem a classe com os atributos para cada produto
class Product:
  def __init__(self, codigo, name, quantity):
    self.codigo = codigo
    self.name = name
    self.quantity = quantity

  def __str__(self):
    return f"{self.codigo} - {self.name} - Quantidade: {self.quantity}"
  #se for colocar preço acrescentar esse item ai em cima - Preço R${self.price:.2f}

  def update_quantity(self, amount):
      if amount >= 0:
          self.quantity = amount
      else:
          print("Não é possível deixar um produto com valor negativo.")

  def increase_quantity(self, amount):
      self.quantity += amount

  def decrease_quantity(self, amount):
      if self.quantity - amount >= 0:
          self.quantity -= amount
      else:
          print("Não é possível diminuir a quantidade abaixo de zero.")

class Inventory:
  def __init__(self):
    self.products = []

  def add_product(self, product):
    self.products.append(product)

  def remove_product(self, product):
    self.products.remove(product)

  def search_product(self, codigo):
    for product in self.products:
      if product.codigo == codigo:
        return product
    return None

  def display_inventory(self):
    print("Lista de Produtos:")
    for product in self.products:
      print(product)

inventory = Inventory()

# Criando uma instância da classe Product
product1 = Product(1, "Camisa", 100)
product2 = Product(2, "Calça", 50)
product3 = Product(3, "Tênis", 30)

inventory.add_product(product1)
inventory.add_product(product2)
inventory.add_product(product3)

#inventory.display_inventory()

#search_result = inventory.search_product("Tênis")
#if search_result:
#  print("Produto encontrado:", search_result)
#else:
#  print("Produto não encontrado.")

#inventory.remove_product(product2)

# Exibindo as informações do produto atualizadas
print(product1)



# Aumentando a quantidade em 50
product1.increase_quantity(50)

# Diminuindo a quantidade em 30
product1.decrease_quantity(30)

# Exibindo as informações do produto atualizadas
print(product1)

inventory.display_inventory()

# Mudando o a quantidade (seria o mesmo para preço)
product1.update_quantity(10)

inventory.display_inventory()
"""
